#para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid #modulo de python para crear un string

from config import connectionBD #conexion a BD

import datetime
import re
import os

from os import remove, path #modulo para remover archivos y para obtener la ruta o direcctorio


import openpyxl #para generar el ecxel

#biblioteca o modulo send_file para forzar la descarga
from flask import send_file

# 1 FUNCION: Funcion que resive del formulario nuevo empleado
def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[°0-9]+', '', dataForm['salario'])
    #convertir salario a INT    
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """INSERT INTO empleados (nombre_empleado, apellidos_empleado, tipo_identidad, n_identidad, fecha_nacimiento, sexo, grupo_rh, email, telefono, profesion, salario, foto_perfil)
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                #Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellidos_empleado'], dataForm['tipo_identidad'],
                           dataForm['n_identidad'], dataForm['fecha_nacimiento'], dataForm['sexo'],
                           dataForm['grupo_rh'], dataForm['email'], dataForm['telefono'],
                           dataForm['profesion'], salario_entero, result_foto_perfil)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert
    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'

# 2 FUNCION: funcion que guarda la imagen del perfil
def procesar_imagen_perfil(foto):
    try:
        #nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]
       
        #creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        #construir la ruta completa de subida de archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, '../static/img/foto_perfil/')

        #validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
        os.chmod(upload_dir, 0o755)

        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile
    except Exception as e:
        print(f"Error al procesar archivo: {e}")
        return []

# 3 FUNCION: Lista de empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """SELECT e.id_empleado, e.nombre_empleado, e.apellidos_empleado, e.salario, e.foto_perfil,
                              CASE WHEN e.sexo = 1 THEN 'Masculino' ELSE 'Femenino' END AS sexo
                              FROM empleados AS e ORDER BY e.id_empleado DESC"""
                cursor.execute(querySQL)
                empleadosBD = cursor.fetchall()
                return empleadosBD
    except Exception as e:
        print(f'Error en la función sql_lista_empleadosBD: {str(e)}')
        return None

# 4 FUNCION: Detalles del empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """SELECT e.id_empleado, e.nombre_empleado, e.apellidos_empleado, e.tipo_identidad,
                              e.n_identidad, e.fecha_nacimiento,
                              CASE WHEN e.sexo = 1 THEN 'Masculino' ELSE 'Femenino' END AS sexo,
                              e.grupo_rh, e.email, e.telefono, e.profesion, e.salario, e.foto_perfil,
                              DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                              FROM empleados AS e WHERE id_empleado = %s ORDER BY e.id_empleado DESC"""
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
                return empleadosBD
    except Exception as e:
        print(f'Error en la función sql_detalles_empleadosBD: {str(e)}')
        return None

# 5 FUNCION: Informe de empleados (reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """SELECT e.id_empleado, e.nombre_empleado, e.apellidos_empleado, e.tipo_identidad,
                              e.n_identidad, e.fecha_nacimiento, e.grupo_rh, e.email, e.telefono,
                              e.profesion, e.salario,
                              DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                              CASE WHEN e.sexo = 1 THEN 'Masculino' ELSE 'Femenino' END AS sexo
                              FROM empleados AS e ORDER BY e.id_empleado DESC"""
                cursor.execute(querySQL)
                empleadosBD = cursor.fetchall()
                return empleadosBD
    except Exception as e:
        print(f'Error en la función empleadosReporte: {str(e)}')
        return None

# 6 FUNCION: Exportar a Excel con empleados
def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    cabeceraExcel = ("Nombre", "Apellido", "Tipo Identidad", "N° Identidad", "Fecha Nacimiento",
                     "Sexo", "Grupo RH", "Email", "Telefono", "Profesion", "Salario", "Fecha de Registro")
    hoja.append(cabeceraExcel)

    formato_moneda_colombiana = '#,##0'

    for registro in dataEmpleados:
        hoja.append((registro['nombre_empleado'], registro['apellidos_empleado'],
                      registro['tipo_identidad'], registro['n_identidad'],
                      registro['fecha_nacimiento'], registro['sexo'],
                      registro['grupo_rh'], registro['email'],
                      registro['telefono'], registro['profesion'],
                      registro['salario'], registro['fecha_registro']))

    for fila_num in range(2, hoja.max_row + 1):
        celda = hoja.cell(row=fila_num, column=7)  # Columna G
        celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/excel"
    ruta_descarga = os.path.join(os.path.dirname(os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
    os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)

# 7 FUNCION: Buscar empleados
def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = """SELECT e.nombre_empleado, e.apellidos_empleado, e.tipo_identidad,
                              e.n_identidad, e.fecha_nacimiento, e.grupo_rh, e.email,
                              e.telefono, e.profesion, e.salario,
                              CASE WHEN e.sexo = 1 THEN 'Masculino' ELSE 'Femenino' END AS sexo
                              FROM empleados AS e WHERE e.nombre_empleado LIKE %s
                              ORDER BY e.id_empleado DESC"""
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {str(e)}")
        return []

# 8 FUNCION: Buscar un empleado
def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = """SELECT e.id_empleado, e.nombre_empleado, e.apellidos_empleado,
                              e.tipo_identidad, e.n_identidad, e.fecha_nacimiento,
                              e.sexo, e.grupo_rh, e.email, e.telefono,
                              e.profesion, e.salario
                              FROM empleados AS e WHERE e.id_empleado = %s LIMIT 1"""
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado
    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {str(e)}")
        return []

# 9 FUNCION: Actualizar un empleado
def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data['form']['nombre_empleado']
                apellidos_empleado = data['form']['apellidos_empleado']
                tipo_identidad = data['form']['tipo_identidad']
                n_identidad = data['form']['n_identidad']
                fecha_nacimiento = data['form']['fecha_nacimiento']
                sexo = data['form']['sexo']
                grupo_rh = data['form']['grupo_rh']
                email = data['form']['email']
                telefono = data['form']['telefono']
                profesion = data['form']['profesion']

                salario_sin_puntos = re.sub('[°0-9]+', '', data['form']['salario'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data['form']['id_empleado']

                if data.files['foto_perfil']:
                    file = data.files['foto_perfil']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """UPDATE empleados SET
                                  nombre_empleado = %s, apellidos_empleado = %s,
                                  tipo_identidad = %s, n_identidad = %s,
                                  fecha_nacimiento = %s, sexo = %s,
                                  grupo_rh = %s, email = %s,
                                  telefono = %s, profesion = %s,
                                  salario = %s, foto_perfil = %s
                                  WHERE id_empleado = %s"""
                    values = (nombre_empleado, apellidos_empleado, tipo_identidad,
                              n_identidad, fecha_nacimiento, sexo, grupo_rh,
                              email, telefono, profesion, salario_empleado,
                              fotoForm, id_empleado)
                else:
                    querySQL = """UPDATE empleados SET
                                  nombre_empleado = %s, apellidos_empleado = %s,
                                  tipo_identidad = %s, n_identidad = %s,
                                  fecha_nacimiento = %s, sexo = %s,
                                  grupo_rh = %s, email = %s,
                                  telefono = %s, profesion = %s,
                                  salario = %s
                                  WHERE id_empleado = %s"""
                    values = (nombre_empleado, apellidos_empleado, tipo_identidad,
                              n_identidad, fecha_nacimiento, sexo, grupo_rh,
                              email, telefono, profesion, salario_empleado, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()
                return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {str(e)}")
        return None

# 10 FUNCION: Lista de usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id_usuario, usuario, email, fecha_creado FROM usuarios"
                cursor.execute(querySQL)
                usuariosBD = cursor.fetchall()
                return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD: {str(e)}")
        return []

# 11 FUNCION: Eliminar empleado
def eliminarEmpleado(id_empleado, foto_perfil):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    basepath = path.dirname(__file__)
                    url_File = path.join(basepath, '../static/img/foto_perfil', foto_perfil)

                    if path.exists(url_File):
                        remove(url_File)

                return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado: {str(e)}")
        return []

# 12 FUNCION: Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM usuarios WHERE id_usuario=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
                return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario: {str(e)}")
        return []     